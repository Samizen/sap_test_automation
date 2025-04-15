import time, re
from playwright.sync_api import Playwright, sync_playwright
from test_playwright.test_login import login_ariba
from test_playwright.config import User_details
from datetime import datetime, timedelta
import shutil
from openpyxl import load_workbook
from pathlib import Path

# Excel Files
original_file_path = Path("excel_test_cases") / "test_edit_pr.xlsx"
copied_file_path = Path("excel_test_cases_updated") / "test_edit_pr_copy.xlsx"
shutil.copyfile(original_file_path, copied_file_path)

# Parameters for test case
USERNAME = User_details.user
PASSWORD = User_details.password
SEARCH_ITEM = "Book Bins - Set of 16"
QUANTITY = "5"
PR_TITLE = "AT-14"
GENERATE_SCRIPT = True
COMPANY_C0DE = "Ariba-Company 1"
COST_CENTER = "Ariba - Quality Assurance"
ACCOUNT_CAT = "ARIBA - Expense Account 2"
# Account Type = Capital (from the dropdown)
PROJECT = "Ariba - Project Two"


# Calculate date 7 days from now
future_date = datetime.now() + timedelta(days=7)
# Format without zero-padding (works on all platforms)
day_name = future_date.strftime("%a")
day = str(int(future_date.strftime("%d")))  # Convert to int and back to string to remove leading zero
month_year = future_date.strftime("%b, %Y")
DELIVERY_DATE = f"{day_name}, {day} {month_year}"


def test_edit_pr():
    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=False)
        context = browser.new_context()
        page = context.new_page()

        # Step 1 - Login to "Advanced B&I"
        login_ariba(page, USERNAME, PASSWORD)
        page.wait_for_load_state("networkidle")
        page.get_by_role("tab", name="Home")

        # Search and Open Requisition 
        page.locator('span.a-no-wrap span.a-srch-portlet-category-dropdown').click()
        page.get_by_role("menuitem", name="Requisition").click()
        time.sleep(3)

        # Search for PR
        page.locator("[id=\"_4opwwd\"]").click()
        page.get_by_role("textbox", name="Title:").fill(PR_TITLE)
        time.sleep(3)
        page.locator('//button[@id="_5z$ioc" and @title="Run this search"]').click()

        # Select PR
        page.locator("[id=\"_4ztpcb\"]").get_by_role("link", name=PR_TITLE).click()
        time.sleep(10)

        # Get PR_number for the requisition for later search 
        element = page.locator(".pageHead.w-page-head")
        full_text = element.text_content().strip()
        full_text = re.sub(r'\s*-\s*', ' - ', full_text)  # normalise spacing around hyphen
        PR_number = full_text.split(" - ")[0].strip()
        print("Extracted PR Number:", PR_number)

        page.get_by_role("button", name="Edit").click()

        # Fill PR details
        # Enter Requisition Title
        page.get_by_role("textbox", name="Title:").fill(PR_TITLE)
        # Enter the remaining data for requisition header - data required (Title, Delivery date - already saved as parameters)
        date_field = page.locator("input.w-txt-img-right-calendar[title*='Enter date']")
        date_field.click()
        date_field.fill(DELIVERY_DATE)

        time.sleep(5)

        # Enter or verify the individual item details - data required (Description*, UOM, Quantity, Price Unit, Discount, Net Amount, Charges, Taxes, Amount, Commodity Code, LblERPCommodityCode, Supplier, Contact, Service Item, Contract)
        page.get_by_role("button", name="Actions").click()
        # requisition_details = extract_receivable_details(page)
        page.locator('[bh="PMI"]', has_text="Edit Details").click()


        # Edit individual detail here:
        # page.locator('#text__jfeesc').click()
        # page.locator('#_jfeesc0').click() # Capital - comment out if no change from Expense
        
        page.locator('div.w-dropdown-item:has-text("Capital")').click()
        page.locator('#_1j0xfc').fill(COST_CENTER) # Cost Center
        page.locator('#_h52ab').fill(COMPANY_C0DE)
        page.locator('#_yj8r6d').fill(ACCOUNT_CAT)
        page.locator('#_4uo7vb').fill(PROJECT)
        # Account Type

        # Test the function
        line_item_details = extract_line_item_details(page)
        accounting_details = extract_accounting_details(page)
        shipping_details = extract_shipping_details(page)      
        # Basic validation    
        
        time.sleep(5)
        page.locator('button[id="_1pa$3d"]').click()

        # Submit PR
        page.get_by_role("button", name="Submit").first.click()
        page.get_by_role("button", name="View Requisition").click()
        page.get_by_role("link", name="Company Logo").click()

        if GENERATE_SCRIPT == True:
            update_excel_template(
                original_file_path="excel_test_cases/test_edit_pr.xlsx",
                copied_file_path="excel_test_cases_updated/test_edit_pr_copy.xlsx",
                SEARCH_ITEM=SEARCH_ITEM,
                QUANTITY=QUANTITY,
                PR_TITLE=PR_TITLE,
                # requisition_details=requisition_details,
                line_item_details=line_item_details,
                accounting_details=accounting_details,
                shipping_details=shipping_details
            )

        # Cleanup
        context.close()
        browser.close()



def extract_line_item_details(page):
    """Enhanced extraction function with better locators and debugging"""
    print("\n=== Starting extraction ===")
    data = {}
    
    # 1. Wait for table to be ready (more flexible selector)
    table = page.locator("table[id='_pm_$pd']")  # Using class instead of ID
    table.wait_for(state="visible")
    print(f"Table visibility: {table.is_visible()}")
    
    # 2. Get all data rows (more reliable selector)
    rows = table.locator('tr:has(td.ffl)').all()  # Rows with label cells
    print(f"Found {len(rows)} potential data rows")
    
    for i, row in enumerate(rows):
        try:
            # 3. Get label
            label = row.locator('td.ffl label').inner_text().strip()
            label = label.replace(':', '').replace(' ', '_')
            print(f"\nProcessing row {i+1} - Label: {label}")
            
            # 4. Get value
            value_cell = row.locator('td.ffp')
            
            # Handle different value types
            if value_cell.locator('input').count():
                value = value_cell.locator('input').input_value()
            elif value_cell.locator('textarea').count():
                value = value_cell.locator('textarea').input_value()
            elif value_cell.locator('a.hoverLink').count():
                value = value_cell.locator('a.hoverLink').inner_text().strip()
            else:
                value = value_cell.inner_text().strip()
            
            value = ' '.join(value.split())
            print(f"Extracted value: {value}")
            
            # Special handling
            if label == 'Discount' and value == 'USD':
                value = ''
            elif label == 'Net_Amount':
                value = value.split('USD')[0].strip()
            
            data[label] = value
            
        except Exception as e:
            print(f"ERROR in row {i+1}: {str(e)}")
            continue
    
    print("\n=== Final Extracted Data ===")
    print(data)
    return data


def extract_accounting_details(page):
    """Extract accounting details from the section and return as dictionary"""
    print("\n=== Starting accounting details extraction ===")
    data = {}
    
    # Wait for accounting section to be visible
    accounting_section = page.locator('div.w-sb-lastSection')
    accounting_section.wait_for(state="visible")
    print(f"Accounting section visible: {accounting_section.is_visible()}")
    
    # Get all accounting tables (there appear to be multiple)
    tables = page.locator('div.w-sb-lastSection table.ftL').all()
    print(f"Found {len(tables)} accounting tables")
    
    for table in tables:
        # Get all rows in the current table
        rows = table.locator('tr:has(td.ffl)').all()
        print(f"Found {len(rows)} rows in current table")
        
        for i, row in enumerate(rows):
            try:
                # Get the label (key)
                label_locator = row.locator('td.ffl label')
                if not label_locator.count():
                    continue
                    
                label = label_locator.inner_text().strip().replace(':', '').replace(' ', '_')
                print(f"\nProcessing row {i+1} - Label: {label}")
                
                # Get the value
                value_cell = row.locator('td.ffp')
                
                # Handle different types of values
                if value_cell.locator('input.w-chInput').count():
                    # For chooser inputs (dropdowns)
                    value = value_cell.locator('input.w-chInput').input_value()
                elif value_cell.locator('div.w-dropdown').count():
                    # For dropdown menus (like Account Type)
                    value = value_cell.locator('div.w-dropdown span.w-dropdown-selected').inner_text().strip()
                elif value_cell.locator('input').count():
                    # Regular input fields
                    value = value_cell.locator('input').input_value()
                elif value_cell.locator('textarea').count():
                    # Textareas
                    value = value_cell.locator('textarea').input_value()
                else:
                    # Plain text
                    value = value_cell.inner_text().strip()
                
                # Clean up the value
                value = ' '.join(value.split())
                print(f"Extracted value: {value}")
                
                data[label] = value
                
            except Exception as e:
                print(f"ERROR in row {i+1}: {str(e)}")
                continue
    
    print("\n=== Final Extracted Accounting Data ===")
    print(data)
    return data


def extract_shipping_details(page):
    """Extract shipping details from the specific table matching the image"""
    print("\n=== Starting targeted shipping extraction ===")
    data = {}
    
    # Wait for the specific shipping table
    shipping_table = page.locator('table.ftL[id="_9p4_ud"]')
    shipping_table.wait_for(state="visible")
    print("Shipping table found and visible")
    
    # Define the specific fields we want to extract
    fields_to_extract = {
        "Ship_To": {
            "locator": 'input.w-chInput[id="_zicgpd"]',
            "type": "input_value"
        },
        "Deliver_To": {
            "locator": 'input.w-txt[id="_tvisgb"]',
            "type": "input_value"
        },
        "Need-by_Date": {
            "locator": 'input.w-txt-img-right-calendar[id="_v471j"]',
            "type": "input_value"
        }
    }
    
    for field_name, config in fields_to_extract.items():
        try:
            element = shipping_table.locator(config["locator"])
            
            if config["type"] == "input_value":
                value = element.input_value()
            else:  # For future expansion if needed
                value = element.inner_text()
                
            # Clean and store the value
            value = ' '.join(value.strip().split())
            data[field_name] = value
            print(f"Extracted {field_name}: {value}")
            
        except Exception as e:
            print(f"Error extracting {field_name}: {str(e)}")
            data[field_name] = None
    
    print("\n=== Final Shipping Data ===")
    print(data)
    return data


def update_excel_template(
    original_file_path,
    copied_file_path,
    SEARCH_ITEM,
    QUANTITY,
    PR_TITLE,
    # requisition_details,
    line_item_details,
    accounting_details,
    shipping_details,
    sheet_name="Steps"
):
    # Convert to Path objects for safety
    original_file_path = Path(original_file_path)
    copied_file_path = Path(copied_file_path)

    # Step 1: Copy the original file
    shutil.copyfile(original_file_path, copied_file_path)

    # Step 2: Load workbook and target sheet
    wb = load_workbook(copied_file_path)
    ws = wb[sheet_name]

    # Step 3: Format dictionaries into multi-line strings
    def dict_to_multiline(data):
        if isinstance(data, dict):
            return "\n".join([f"{key}: {value}" for key, value in data.items()])
        return str(data)

    # Step 4: Write values to specified cells
    ws['G6'] = PR_TITLE
    # ws['G11'] = dict_to_multiline(requisition_details)
    ws['G8'] = dict_to_multiline(line_item_details)
    ws['G11'] = dict_to_multiline(accounting_details)
    ws['G12'] = dict_to_multiline(shipping_details)

    # Step 5: Save the modified file
    wb.save(copied_file_path)
    print(f"Excel file updated and saved to: {copied_file_path}")



# Home ma 

        