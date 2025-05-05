from playwright.sync_api import sync_playwright
from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
import time, re
from test_playwright.test_login import login_ariba
from test_playwright.config import User_details
from datetime import datetime
from html import unescape
from openpyxl import load_workbook
from pathlib import Path
import shutil
from openpyxl.styles import Alignment

original_file_path = Path("excel_test_cases") / "test_create_invoice.xlsx"
copied_file_path = Path("excel_test_cases_updated") / "test_create_invoice_copy.xlsx"
shutil.copyfile(original_file_path, copied_file_path)

# Parameters for test case
URL = "https://s1.au.cloud.ariba.com/Buyer/Main/ad/loginPage/SSOActions?awsso_cc=cmVhbG06VUZKRlRVbExRVlJKTFVSRlRVOUVVMEZRVUMweExWUT07YXdzc29fcnU6YUhSMGNITTZMeTl6TVM1aGRTNWpiRzkxWkM1aGNtbGlZUzVqYjIwdlFuVjVaWEl2VFdGcGJpOWhaQzlrWldaaGRXeDBMMFJwY21WamRFRmpkR2x2Ymo5eVpXRnNiVDFRVWtWTlNVdEJWRWt0UkVWTlQwUlRRVkJRTFRFdFZBPT07YXdzc29fbHU6YUhSMGNITTZMeTl6TVM1aGRTNWpiRzkxWkM1aGNtbGlZUzVqYjIwdlFuVjVaWEl2VFdGcGJpOWhaQzlqYkdsbGJuUk1iMmR2ZFhRdlUxTlBRV04wYVc5dWN3PT07YXdzc29fYXA6UW5WNVpYST07YXdzc29fYXJpZDpNVGMwTVRJME9EVXhNak16Tmc9PTthd3Nzb19rdTphSFIwY0hNNkx5OXpNUzVoZFM1amJHOTFaQzVoY21saVlTNWpiMjB2UW5WNVpYSXZUV0ZwYmk5aFpDOWpiR2xsYm5STFpXVndRV3hwZG1VdlUxTlBRV04wYVc5dWN3PT07YXdzc29fZmw6TVE9PQ%3D%3D%3ATymCh4RVvNfhYbiPc0XDQwVRWls%3D&awsso_ap=Buyer&realm=PREMIKATI-DEMODSAPP-1-T&awsr=true#b0"
# Parameters for test case
USERNAME = User_details.user
PASSWORD = User_details.password
# PO_ID = "PO191"  # Replace with your PO ID
INVOICE_ID = "12344545452321"  # Replace with your invoice ID
INVOICE_DATE = datetime.today().strftime("%a, %d %b, %Y")
# REMIT_ADDRESS = "1000005900"  # Replace with your remit address
SEARCH_QUERY = "AT-30"  # Replace with your search query

def test_invoice_legacy():
    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=False)
        context = browser.new_context()
        page = context.new_page()

        # Navigate to the login page
        page.goto(URL)

        # Login
        login_ariba(page, USERNAME, PASSWORD)
        time.sleep(5)

        # Navigate to Home and click Create
        page.get_by_role("tab", name="Home").click()
        time.sleep(5)
        page.get_by_role("button", name="Create").click()
        time.sleep(3)

        # Go to Invoice
        page.get_by_role("menuitem", name="Invoice").click()
        time.sleep(3)

        # Select PO-Based
        page.locator("label").filter(has_text="PO-Based").locator("label").click()
        time.sleep(3)

        # Select PO ID
        page.locator("[id=\"_aeqrid\"]").get_by_role("link").click()
        page.get_by_role("option", name="Search more").click()
        page.get_by_role("combobox", name="Search for a specific value").locator("span").nth(1).click()
        page.get_by_role("option", name="Title").click()
        page.get_by_role("textbox", name="Search for a specific value").click()
        page.get_by_role("textbox", name="Search for a specific value").click()
        page.get_by_role("textbox", name="Search for a specific value").fill(SEARCH_QUERY)
        page.get_by_title("Search for a specific value").click()

        # row = page.locator('tr[dr="1"]').filter(has_text=SEARCH_QUERY)
        # # Click the checkbox inside the row
        # row.locator('input[type="checkbox"]').first.click()

        row = page.locator('tr[dr="1"]').filter(has_text=SEARCH_QUERY)

        # Extract the PO ID text dynamically from the second column <a> tag
        po_id_element = row.locator("td").nth(1).locator("a span")
        PO_ID = po_id_element.inner_text()

        row_locator = page.locator(f"tr:has-text('{PO_ID}')")

        row_locator.locator("label.w-chk").first.click()
        time.sleep(2)
        page.get_by_role("button", name="Done").click()


        time.sleep(4)

        # Fill Invoice Details
        page.get_by_role("textbox", name="Supplier Invoice #:").click()
        page.get_by_role("textbox", name="Supplier Invoice #:").fill(INVOICE_ID)
        time.sleep(3)

        # Fill Invoice Date
        page.get_by_role("textbox", name="Invoice Date:").fill(INVOICE_DATE)
        time.sleep(3)

        # Fill Remit Address
        # page.locator('input[id="_a_inqd"][type="text"][role="combobox"]').fill(REMIT_ADDRESS)

        purchase_order_data = extract_purchase_order_details(page)
        print(purchase_order_data)
        payment_terms_data = extract_payment_terms_details(page)
        print(payment_terms_data)
        line_items_data = extract_line_items_details(page)
        print(line_items_data)
        update_excel_template_with_po_terms(
            original_file_path,
            copied_file_path,
            purchase_order_data,
            payment_terms_data,
            line_items_data
        )

        page.locator("[id=\"_ihiawc\"] > td").first.click()
        page.get_by_role("button", name="Edit").click()
        page.get_by_role("textbox", name="Reference Date:").fill(INVOICE_DATE)
        page.get_by_role("button", name="Validate and Exit").first.click()
        #Submit the invoice
        time.sleep(3)
        page.get_by_role("button", name="Submit").first.click()

        # Cleanup
        context.close()
        browser.close()


def extract_purchase_order_details(page):
    """Enhanced extraction function for Purchase Order and Type details"""
    print("\n=== Starting extraction ===")
    data = {}
    
    # 1. Wait for table to be ready (using a more flexible selector)
    table = page.locator("table[id='_sdur1']")  # Select the table by ID or class
    table.wait_for(state="visible")
    print(f"Table visibility: {table.is_visible()}")
    
    # 2. Extract label and value pairs
    rows = table.locator('tr:has(td.ffl)').all()  # Get rows containing labels
    print(f"Found {len(rows)} potential data rows")
    
    for i, row in enumerate(rows):
        try:
            # 3. Get the label text
            label = row.locator('td.ffl label').inner_text().strip()
            label = label.replace(':', '').replace(' ', '_')  # Normalize label
            print(f"\nProcessing row {i+1} - Label: {label}")
            
            # 4. Get the value for the label
            value_cell = row.locator('td.ffp')
            
            # Handle different input types (radio buttons, text input, etc.)
            if value_cell.locator('input[type="radio"]').count():
                # Extract selected radio button option
                selected_radio = value_cell.locator('input[type="radio"]:checked')
                value = selected_radio.locator('~label').inner_text().strip() if selected_radio.count() else ''
            elif value_cell.locator('input[type="text"]').count():
                value = value_cell.locator('input[type="text"]').input_value().strip()
            elif value_cell.locator('textarea').count():
                value = value_cell.locator('textarea').input_value().strip()
            elif value_cell.locator('a').count():
                value = value_cell.locator('a').inner_text().strip()  # For links or selectors
            else:
                value = value_cell.inner_text().strip()  # Default for regular text
            
            value = ' '.join(value.split())  # Normalize spaces
            print(f"Extracted value: {value}")
            
            # Special handling for certain labels (if needed)
            if label == 'Purchase_Orders':
                value = value.split('PO')[0].strip()  # Custom handling if necessary
            
            # Add extracted data to the dictionary
            data[label] = value
        
        except Exception as e:
            print(f"ERROR in row {i+1}: {str(e)}")
            continue
    
    print("\n=== Final Extracted Data ===")
    print(data)
    return data


def extract_payment_terms_details(page):
    """Extracts details from the '_880gn' table including fields like Payment Terms and Ship From"""
    print("\n=== Starting extraction for additional details ===")
    data = {}

    # 1. Wait for the table to become visible
    table = page.locator("table#_880gn")
    table.wait_for(state="visible")
    print(f"Table visibility: {table.is_visible()}")

    # 2. Get all rows with 'td.ffl' (label) class
    rows = table.locator('tr:has(td.ffl)').all()
    print(f"Found {len(rows)} rows to process")

    for i, row in enumerate(rows):
        try:
            label_locator = row.locator('td.ffl label')
            label = label_locator.inner_text().strip().replace(':', '').replace(' ', '_')
            print(f"\nProcessing row {i+1} - Label: {label}")

            value_cell = row.locator('td.ffp')
            value = ""

            # Try getting combobox by role
            if value_cell.get_by_role("combobox").count() == 1:
                value = value_cell.get_by_role("combobox").input_value().strip()
            # Then try textarea
            elif value_cell.locator('textarea').count() == 1:
                value = value_cell.locator('textarea').input_value().strip()
            # Then try anchor
            elif value_cell.locator('a').count() >= 1:
                value = value_cell.locator('a').first.inner_text().strip()
            # Fallback to plain text
            else:
                value = value_cell.inner_text().strip()

            value = ' '.join(value.split())  # Normalize whitespace
            print(f"Extracted value: {value}")

            data[label] = value

        except Exception as e:
            print(f"ERROR in row {i+1}: {str(e)}")
            continue

    print("\n=== Final Extracted Additional Data ===")
    print(data)
    return data


def extract_line_items_details(page):
    """Extracts line item details from the '_coc8ad' table (Line Items table)."""
    print("\n=== Starting extraction for line items ===")
    data = []

    try:
        # Wait for the table to be visible
        line_items_table = page.locator('table#_coc8ad')
        line_items_table.wait_for(state='visible', timeout=10000)
        print(f"Line Items Table visibility: {line_items_table.is_visible()}")

        # Locate rows using the correct selector
        rows = line_items_table.locator('tr[id^="_ihiawc"]').all()  # Using the `id` prefix for row elements
        print(f"Found {len(rows)} line item rows")

        if len(rows) == 0:
            print("No line item rows found in the table. Please check the table structure.")
            return data  # Return empty list if no rows found

        for i, row in enumerate(rows):
            try:
                print(f"\nProcessing row {i+1}")

                # Extract values with fallback handling for optional inputs
                line_data = {
                    "No": row.locator('td#_jdsrc').inner_text().strip(),
                    "Description": row.locator('td#_ms4chc').inner_text().strip(),
                    "Order_ID": row.locator('td#_pk2kzd a span').inner_text().strip(),
                    "Qty": row.locator('input#_bwol9b').input_value().strip(),
                    "Unit": row.locator('td#_2wpqf a').inner_text().strip(),
                    "Price": row.locator('td#_o8sm2d').inner_text().strip(),
                    "Amount": row.locator('td#_hngptb').inner_text().strip(),
                    "Discount": row.locator('input#_gj0$q').input_value().strip() if row.locator('input#_gj0$q').count() > 0 else "",
                    "Gross_Amount": row.locator('td#_4gloub').inner_text().strip(),
                }

                # Normalize spacing and clean up values
                line_data = {k: ' '.join(v.split()) if isinstance(v, str) else v for k, v in line_data.items()}
                print(f"Extracted line item: {line_data}")
                data.append(line_data)

            except Exception as e:
                print(f"ERROR in row {i+1}: {str(e)}")
                continue

    except Exception as e:
        print(f"General extraction error: {str(e)}")
        # Take screenshot for debugging
        page.screenshot(path="line_items_error.png")
    
    print("\n=== Final Extracted Line Items Data ===")
    print(data)
    return data



def update_excel_template_with_po_terms(
    original_file_path,
    copied_file_path,
    purchase_order_data,
    payment_terms_data,
    line_items_data
):
    # Copy the base template
    shutil.copyfile(original_file_path, copied_file_path)

    # Load the copied Excel workbook and active sheet
    wb = load_workbook(copied_file_path)
    ws = wb.active

    # Combine the two dictionaries
    combined_dict = {**purchase_order_data, **payment_terms_data}

    # Format dictionary data as a string with new lines after each element
    combined_dict_str = '\n'.join(f'{k}: {v}' for k, v in combined_dict.items())

    # Format line items data as a string with new lines for each line item
    line_items_str = '\n'.join(
        [', '.join(f'{k}: {v}' for k, v in item.items()) for item in line_items_data]
    )

    # Write to H7 and H8
    ws['H7'] = combined_dict_str
    ws['H8'] = line_items_str

    # Apply text wrapping to H7 and H8 to show new lines properly
    ws['H7'].alignment = Alignment(wrap_text=True)
    ws['H8'].alignment = Alignment(wrap_text=True)

    # Save the workbook
    wb.save(copied_file_path)