import os
import pandas as pd
# # from excel_name_change.config import vars_to_change 

vars_to_change = {
    "Complete":"__completed__",
}

# class Excel_Name:

#     def __init__(self, files_dir=""):
#         self.files_dir = files_dir
#         self.excel_files = os.listdir(self.files_dir)
    
#     def read_through_df(self,df):
#         for var in vars_to_change:
#             print(f"Changing {var} to {vars_to_change[var]}")
#             df.replace(var, vars_to_change[var], inplace=True)
#         return df

#     def change_name(self, file):
#         file_path = os.path.join(self.files_dir, file)
#         dfs = pd.read_excel(file_path, engine="openpyxl", sheet_name=None)
#         for sheet, df in dfs.items():
#             print(f"\nSheet: {sheet}")
#             dfs[sheet] = self.read_through_df(df)

#             # Save back to Excel with multiple sheets
#             output_path = "excel_name_change/output/output.xlsx"
#             os.makedirs(os.path.dirname(output_path), exist_ok=True)  # Ensure directory exists

#             with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
#                 for sheet_name, df in dfs.items():
#                     df.to_excel(writer, sheet_name=sheet_name, index=False)

#             print(f"Updated file saved at: {output_path}")

#     def change_on_all(self):
#         print(self.excel_files[0])
#         # for file in self.excel_files:
#         #     if file.endswith(".xlsx"):
#         #         os.rename(os.path.join(self.files_dir, file), os.path.join(self.files_dir, file[:-5] + ".xls"))

#         self.change_name(self.excel_files[0])
#         return


# Excel_Name(files_dir="excel_name_change/excel_files/").change_on_all()


from openpyxl import load_workbook

class Excel_Name:
    def __init__(self, files_dir=""):
        self.files_dir = files_dir
        self.excel_files = os.listdir(self.files_dir)
    
    def read_through_df(self, df:pd.DataFrame):
        for var in vars_to_change:
            print(f"Changing {var} to {vars_to_change[var]}")
            df.replace(var, vars_to_change[var], inplace=True)
        return df

    def change_name(self, file):
        file_path = os.path.join(self.files_dir, file)
        wb = load_workbook(file_path)

        for sheet in wb.sheetnames:
            ws = wb[sheet]
            df = pd.DataFrame(ws.values)

            # Replace values in DataFrame (data-only)
            df = self.read_through_df(df)

            # Write the modified DataFrame back to the sheet
            for row_index, row in enumerate(df.values, start=1):
                for col_index, value in enumerate(row, start=1):
                    ws.cell(row=row_index, column=col_index, value=value)

        output_path = "excel_name_change/output/output.xlsx"
        os.makedirs(os.path.dirname(output_path), exist_ok=True)

        # Save workbook with original formatting
        wb.save(output_path)
        print(f"Updated file saved at: {output_path}")

    def change_on_all(self):
        self.change_name(self.excel_files[0])

Excel_Name(files_dir="excel_name_change/excel_files/").change_on_all()
